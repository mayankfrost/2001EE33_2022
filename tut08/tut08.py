from datetime import datetime
import re
import csv
from openpyxl import Workbook

start_time = datetime.now()


def cleanse(name: str):
    if name[-1] == ')':  # remove the (c) and (w) from player names
        name = name[:-3]
    return name


def play_inning(inn: list[str], batsmen: list[str], bowlers: list[str], row, sheet):
    # common statistics for bowlers and batsmen
    balls = {}
    runs = {}

    # statistics for batsmen
    fours = {}
    sixes = {}
    status = {}

    # statistics for bowlers
    wickets = {}
    nb = {}
    maiden = {}
    wd = {}
    streak = {}

    # overall statistics
    byes = lbs = wide = nbs = pens = 0
    tot_runs = tot_wkts = tot_balls = 0

    fow = []  # fall of wickets

    for player in batsmen:
        balls[player] = runs[player] = fours[player] = sixes[player] = 0
        status[player] = 'B'

    for player in bowlers:
        balls[player] = runs[player] = wickets[player] = nb[player] = maiden[player] = wd[player] = streak[player] = 0

    alt_names = {}
    players = batsmen + bowlers

    for player in players:
        alt_names[cleanse(player)] = player
        for names in re.split(' ', player):  # capturing the first and last names and mapping it to the full names
            alt_names[cleanse(names)] = player

    for step in inn:
        if step == '\n':  # if it's an empty line, simply skip it
            continue

        bowler = alt_names[step[step.find(' ') + 1:step.find('to') - 1]]
        batsman = alt_names[step[step.find('to') + 3:step.find(',')]]
        pos = step.find(',') + 2
        run_char = step[pos]
        run = 0

        status[batsman] = 'not out'

        # examining the characters and figuring out what is happening during a particular ball

        if run_char == 'n':
            streak[bowler] += 1

        elif run_char == 'o':
            streak[bowler] += 1
            wickets[bowler] += 1
            tot_wkts += 1

            fow.append(
                str(tot_runs) + '-' + str(tot_wkts) + ' (' + cleanse(batsman) + ', ' + step[:step.find('.') + 2] + ')')

            if step[pos + 4] == 'L':
                status[batsman] = 'lbw b ' + cleanse(bowler)
            elif step[pos + 4] == 'C':
                catcher = alt_names[step[pos + 14:step.find('!')]]
                status[batsman] = 'c ' + cleanse(catcher) + ' b ' + cleanse(bowler)
            else:
                status[batsman] = 'b ' + cleanse(bowler)
        else:
            streak[bowler] = 0
            if run_char == 'F':
                run = 4
                fours[batsman] += 1
            elif run_char == 'S':
                run = 6
                sixes[batsman] += 1

            elif run_char == 'w':  # in case of wide
                runs[bowler] += 1
                wd[bowler] += 1
                wide += 1
                tot_runs += 1
                continue

            elif run_char == 'b' or run_char == 'l':  # in case of byes of lbs

                runc2 = step[pos + (6 if run_char == 'b' else 10)]
                run2 = 0
                if runc2 == 'F':
                    run2 = 4
                else:
                    run2 = int(runc2)
                runs[bowler] += run2
                tot_runs += run2
                if run_char == 'b':
                    byes += run2
                else:
                    lbs += run2

            else:
                if step[pos + 2] == 'w':
                    run = int(run_char)
                    runs[bowler] += run
                    wd[bowler] += run
                    wide += run
                    tot_runs += run
                    continue

                run += int(run_char)

        # making appropriate adjustments to our statistics

        runs[batsman] += run
        runs[bowler] += run
        tot_runs += run

        balls[batsman] += 1
        balls[bowler] += 1
        tot_balls += 1

        if step[2] == '6' and streak[bowler] >= 6:
            maiden[bowler] += 1

    # making a record for the batting table

    sheet.cell(row=row, column=1).value = 'Batter'
    sheet.cell(row=row, column=3).value = 'R'
    sheet.cell(row=row, column=4).value = 'B'
    sheet.cell(row=row, column=5).value = '4s'
    sheet.cell(row=row, column=6).value = '6s'
    sheet.cell(row=row, column=7).value = 'SR'
    row += 2

    for batter in batsmen:
        if status[batter] == 'B':
            continue

        sheet.cell(row=row, column=1).value = batter
        sheet.cell(row=row, column=2).value = status[batter]
        sheet.cell(row=row, column=3).value = runs[batter]
        sheet.cell(row=row, column=4).value = balls[batter]
        sheet.cell(row=row, column=5).value = fours[batter]
        sheet.cell(row=row, column=6).value = sixes[batter]
        sheet.cell(row=row, column=7).value = round(runs[batter] * 100 / balls[batter], 2) if balls[batter] > 0 else 0

        row += 1

    row += 1
    sheet.cell(row=row, column=1).value = 'Extras'
    sheet.cell(row=row, column=2).value = '{} (b {}, lb {}, w {}, nb {}, p {})'.format(byes + lbs + wide + nbs + pens,
                                                                                       byes, lbs, wide, nbs, pens)
    row += 1
    sheet.cell(row=row, column=1).value = 'Total'
    sheet.cell(row=row, column=2).value = tot_runs
    sheet.cell(row=row, column=3).value = '(' + str(tot_wkts) + ' wkts, ' + str(tot_balls // 6) + '.' + str(
        tot_balls % 6) + ' Ov)'
    row += 1

    sheet.cell(row=row, column=1).value = 'Did not Bat:'
    c = 2
    for b in batsmen:
        if status[b] == 'B':
            sheet.cell(row=row, column=c).value = cleanse(b)
            c += 1
    row += 1

    sheet.cell(row=row, column=1).value = 'Fall of Wickets'
    row += 1
    c = 1
    for f in fow:
        sheet.cell(row=row, column=c).value = f
        c += 1
    row += 2

    # making a record for the bowling table

    sheet.cell(row=row, column=1).value = 'Bowler'
    sheet.cell(row=row, column=2).value = 'O'
    sheet.cell(row=row, column=3).value = 'M'
    sheet.cell(row=row, column=4).value = 'R'
    sheet.cell(row=row, column=5).value = 'W'
    sheet.cell(row=row, column=6).value = 'NB'
    sheet.cell(row=row, column=7).value = 'WD'
    sheet.cell(row=row, column=8).value = 'ECO'
    row += 1

    for b in bowlers:
        if balls[b] == 0:
            continue

        sheet.cell(row=row, column=1).value = b
        sheet.cell(row=row, column=2).value = str(balls[b] // 6) + ('.' + str(balls[b] % 6) if balls[b] % 6 else '')
        sheet.cell(row=row, column=3).value = maiden[b]
        sheet.cell(row=row, column=4).value = runs[b]
        sheet.cell(row=row, column=5).value = wickets[b]
        sheet.cell(row=row, column=6).value = nb[b]
        sheet.cell(row=row, column=7).value = wd[b]
        sheet.cell(row=row, column=8).value = round(runs[b] * 6 / balls[b], 2)
        row += 1

    return row + 2


def scorecard():
    file_team = open("teams.txt", "r")
    teams = file_team.readlines()

    # replacing ':' with ',' for easier detection of players
    pak_team = re.sub(':', ',', teams[0][:-1])
    ind_team = re.sub(':', ',', teams[2][:-1])

    pak_players = re.split(', ', pak_team)
    del pak_players[0]  # removing the first part of the split since it's not a player
    ind_players = re.split(', ', ind_team)
    del ind_players[0]  # removing the first part of the split since it's not a player

    pak_inn_f = open('pak_inns1.txt', 'r')
    pak_inn = pak_inn_f.readlines()
    ind_inn_f = open('india_inns2.txt', 'r')
    ind_inn = ind_inn_f.readlines()

    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'Pakistan Innings'
    row = play_inning(pak_inn, pak_players, ind_players, 2, sheet)  # run Pakistan's innings
    sheet.cell(row=row, column=1).value = 'India Innings'
    play_inning(ind_inn, ind_players, pak_players, row + 1, sheet)  # run India's innings

    op = csv.writer(open('Scorecard.csv', 'w', newline=''))
    for r in sheet.rows:  # writing openpyxl rows onto a csv file
        op.writerow([cell.value for cell in r])

    # closing the file objects
    file_team.close()
    pak_inn_f.close()
    ind_inn_f.close()


scorecard()

# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
