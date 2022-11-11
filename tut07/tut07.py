import pandas as pd
import os

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side

from datetime import datetime

start_time = datetime.now()



def pos_of_oc(oc):
    return abs(oc) * 2 - (oc > 0)


def find_oc(u, v, w):
    oc = 1
    if w < 0:
        oc = -1

    if u < 0:
        if v < 0:
            oc *= 3
        else:
            oc *= 2

    elif v < 0:
        oc *= 4

    return oc


def filewise_analysis(file_name, mod=5000):

    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction",
                              "2": "External Ejection", "-2": "Internal Ejection", "3": "External inward interaction",
                              "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
    vals = [1, -1, 2, -2, 3, -3, 4, -4]  # values for different octants

    data = pd.read_excel(os.path.join('input', file_name))

    total = data['U'].shape[0]  # total number of coordinates

    u_avg = 0
    v_avg = 0
    w_avg = 0

    try:
        for i in data['U']:
            u_avg += i
    except TypeError:
        print("Column of U has some data type besides int and float")

    try:
        for i in data['V']:
            v_avg += i
    except TypeError:
        print("Column of V has some data type besides int and float")

    try:
        for i in data['W']:
            w_avg += i
    except TypeError:
        print("Column of W has some data type besides int and float")

    u_avg = round(u_avg / total, 3)  # finding out the averages
    v_avg = round(v_avg / total, 3)
    w_avg = round(w_avg / total, 3)

    ua = [u_avg]
    ua.extend([''] * (total - 1))
    va = [v_avg]
    va.extend([''] * (total - 1))
    wa = [w_avg]
    wa.extend([''] * (total - 1))

    data['U Avg'] = ua
    data['V Avg'] = va
    data['W Avg'] = wa

    uu = []
    vv = []
    ww = []

    for i in data['U']:
        uu.append(i - u_avg)

    for i in data['V']:
        vv.append(i - v_avg)

    for i in data['W']:
        ww.append(i - w_avg)

    data['U\'=U - U avg'] = uu
    data['V\'=V - V avg'] = vv
    data['W\'=W - W avg'] = ww

    octant = []

    cnt = {}
    rank = {}
    scores = {}

    mrows = (total + mod - 1) // mod  # number of rows for each range of count

    trans_cnt = {}
    trans_id = [''] * total

    unit = 13  # rows for one transition count
    buffer = 11

    for i in vals:  # generating the tables and filling the initial values with 1
        trans_cnt[i] = [0] * (buffer + unit * mrows)
        trans_cnt[i][0] = ''
        trans_cnt[i][1] = str(i) if i < 0 else '+' + str(i)
        trans_cnt[i][buffer - 1] = ''

        for j in range(mrows):
            for k in range(4):
                trans_cnt[i][buffer + j * unit + k] = ''
            trans_cnt[i][buffer + j * unit + 4] = str(i) if i < 0 else '+' + str(i)

    for i in vals:
        cnt[i] = [0] * (mrows + 2)
        cnt[i][1] = ''

        rank[i] = [0] * (mrows + 2)
        rank[i][1] = ''

        scores[i] = 0

    prev = 0

    for i in range(total):

        # determining which octant it belongs to
        oc = find_oc(uu[i], vv[i], ww[i])

        octant.append(str(oc) if oc < 0 else '+' + str(oc))
        cnt[oc][0] += 1
        cnt[oc][2 + i // mod] += 1

        if i > 0:
            trans_cnt[oc][1 + pos_of_oc(prev)] += 1  # overall transition count

            mrange = (i - 1) // mod  # determing which range this current octant belongs to
            trans_cnt[oc][buffer + mrange * unit + 4 + pos_of_oc(prev)] += 1  # overall transition count

        prev = oc

    for i in vals:
        cnt[i].extend([''] * (total - len(cnt[i])))
        rank[i].extend([''] * (total - len(rank[i])))
        trans_cnt[i].extend([''] * (total - len(rank[i])))

    data['Octant'] = octant
    data['#1'] = [''] * total

    almost_emp_col = ['', 'User Input']  # creating an almost empty column which will have "User Input" in 2nd row
    almost_emp_col.extend([''] * (total - 2))
    data['#2'] = almost_emp_col

    id = ['Overall Count', 'Mod ' + str(mod)]

    for i in range(mrows):
        id.append(str(i * mod) + '-' + str(min(total - 1, (i + 1) * mod - 1)))
    id.extend([''] * (total - len(id)))

    data['Octant ID'] = id

    order = []
    for i in vals:
        order.append([cnt[i][0], i])
    order.sort(reverse=True)  # arranging the octant values in order of their occurences at different timestamps
    to_yellow = []

    for i in vals:
        data[str(i) if i < 0 else '+' + str(i)] = cnt[i]

    for i in range(8):
        rank[order[i][1]][0] = i + 1  # assigning ranks according to the order

    to_yellow.append([2, 22 + pos_of_oc(order[0][1])])

    r1_id = [order[0][1], '']

    for i in range(mrows):
        order.clear()
        for j in vals:
            order.append([cnt[j][i + 2], j])
        order.sort(reverse=True)  # arranging the octant values in order of their occurences at different timestamps

        r1_id.append(order[0][1])
        scores[order[0][1]] += 1  # incrementing the score of rank 1.
        to_yellow.append([i + 4, 22 + pos_of_oc(order[0][1])])

        for j in range(8):
            rank[order[j][1]][i + 2] = j + 1  # assigning ranks according to the order

    r1_on = []
    for i in r1_id:
        r1_on.append(octant_name_id_mapping[str(i)] if i != '' else '')

    r1_id.extend([''] * (total - len(r1_id)))
    r1_on.extend([''] * (total - len(r1_on)))

    row = mrows + 5
    rank[4][row] = 'Octant ID'
    rank[-4][row] = 'Octant Name'
    r1_id[row] = 'Count of Rank 1 Mod Values'

    for i in range(8):  # writing the scores of each octant value
        row = mrows + 6 + i
        rank[4][row] = vals[i]
        rank[-4][row] = octant_name_id_mapping[str(vals[i])]
        r1_id[row] = scores[vals[i]]

    for i in vals:
        data[('+' if i > 0 else '') + str(i)] = cnt[i]
    for i in vals:
        data['Rank Octant ' + ('+' if i > 0 else '') + str(i)] = rank[i]

    data['Rank1 Octant ID'] = r1_id
    data['Rank1 Octant Name'] = r1_on

    data['#3'] = [''] * total

    emp_col = [''] * total
    data['#4'] = emp_col

    trans_id[1] = 'Octant #'
    trans_cnt[1][0] = 'To'
    c = 2
    for j in vals:
        trans_id[c] = str(j) if j < 0 else '+' + str(j)
        c += 1

    from_col = [''] * total  # column for "from"s
    from_col[2] = 'From'

    for i in range(mrows):
        # generating the tables for transitions

        trans_id[buffer + i * unit + 2] = 'Mod Transition Count'
        trans_id[buffer + i * unit + 4] = 'Octant #'
        from_col[buffer + i * unit + 5] = 'From'
        trans_cnt[1][buffer + i * unit + 3] = 'To'

        c = 5
        trans_id[buffer + i * unit + 3] = str(i * mod) + '-' + str(min(total - 1, (i + 1) * mod - 1))
        for j in vals:
            trans_id[buffer + i * unit + c] = str(j) if j < 0 else '+' + str(j)
            c += 1

    data['#5'] = from_col
    data['Overall Transition Count'] = trans_id

    c = 11

    for i in vals:
        trans_cnt[i].extend([''] * (total - len(trans_cnt[i])))
        data['#' + str(c)] = trans_cnt[i]
        c += 1

    data['#30'] = [''] * total

    sub_len = {}
    sub_cnt = {}
    ranges = {}

    for i in vals:
        sub_cnt[i] = 0  # count of
        sub_len[i] = 0  # length of longest subsequence
        ranges[i] = []

    length = 0  # length of the current subsequence
    prev = 0  # previous octant

    for i in range(total):

        oc = find_oc(uu[i], vv[i], ww[i])

        if oc == prev:  # adjusting length of the current running subsequence
            length += 1
        else:
            length = 1

        if sub_len[oc] == length:  # using the current length to update the subsequence length, count and time ranges
            sub_cnt[oc] += 1
            ranges[oc].append([data['T'][i - length + 1], data['T'][i]])

        elif sub_len[oc] < length:
            sub_cnt[oc] = 1
            ranges[oc] = [[data['T'][i - length + 1], data['T'][i]]]
            sub_len[oc] = length

        prev = oc

    oc_col = ['', 'Octant ##']
    len_col = ['', 'Longest Subsequence Length']
    cnt_col = ['', 'Count']
    for i in vals:
        oc_col.append(str(i) if i < 0 else '+' + str(i))
        len_col.append(sub_len[i])
        cnt_col.append(sub_cnt[i])

    oc_col.extend([''] * (total - len(oc_col)))
    len_col.extend([''] * (total - len(len_col)))
    cnt_col.extend([''] * (total - len(cnt_col)))

    data['Longest Subsquence Length'] = oc_col
    data['#31'] = len_col
    data['#32'] = cnt_col

    data['#33'] = [''] * total

    # Reusing the 3 lists since we don't their older values have already been copied to the pandas dataframe

    oc_col = ['', 'Octant ###']
    len_col = ['', 'Longest Subsequence Length']
    cnt_col = ['', 'Count']

    for i in vals:
        oc_col.append(str(i) if i < 0 else '+' + str(i))
        oc_col.append('T')
        oc_col.extend([''] * len(ranges[i]))  # leaving an appropriate amount of empty lines

        len_col.append(sub_len[i])
        len_col.append('From')

        cnt_col.append(sub_cnt[i])
        cnt_col.append('To')

        for x, y in ranges[i]:
            len_col.append(x)
            cnt_col.append(y)

    oc_col.extend([''] * (total - len(oc_col)))
    len_col.extend([''] * (total - len(len_col)))
    cnt_col.extend([''] * (total - len(cnt_col)))

    data['Longest Subsequence Length with Range'] = oc_col
    data['#35'] = len_col
    data['#36'] = cnt_col

    # Importing the pandas dataframe to openpyxl workbook for stylizing it

    wb = Workbook()
    sheet = wb.active

    for r in dataframe_to_rows(data, index=False, header=True):
        sheet.append(r)

    # highlighting required cells to yellow

    for r, c in to_yellow:
        sheet.cell(row=r, column=c).fill = PatternFill(start_color="00FFFF00", fill_type="solid")

    # empty headers were given a dummy name starting with '#' for convenience in pandas
    # those headers are made empty now

    c = 1
    for cell in sheet.iter_cols(min_row=1, max_row=1):
        cell = cell[0]
        if cell.value[0] == '#':
            sheet.cell(row=1, column=c).value = ''
        c += 1

    # defining our style of border

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # appending the cells which require a border to a list

    give_border = []

    for i in range(14, 33):
        for j in range(1, mrows + 4):
            give_border.append([j, i])

    for i in range(29, 32):
        for j in range(mrows + 7, mrows + 16):
            give_border.append([j, i])

    for i in range(36, 45):
        for j in range(3, 12):
            give_border.append([j, i])
        for j in range(mrows):
            for k in range(9):
                give_border.append([buffer + j * unit + 6 + k, i])

    for i in range(46, 49):
        for j in range(3, 12):
            give_border.append([j, i])

    for i in range(50, 53):
        sum = 17
        for j in sub_cnt.values():
            sum += j

        for j in range(3, sum + 3):
            give_border.append([j, i])

    # giving borders to the required cells

    for r, c in give_border:
        sheet.cell(row=r, column=c).border = border

    if not os.path.isdir('output'):  # Make a new directory 'output' if it does not already exist
        os.mkdir('output')

    # saving the final file
    wb.save(os.path.join('output', file_name[:-5] + ' cm_vel_octant_analysis_mod_' + str(mod) + '.xlsx'))
    print('Output for ' + file_name + 'has been created')


def octant_analysis(mod):
    for file in os.listdir('input'):  # for each file in input
        filewise_analysis(file, mod)


mod = 5000  # changeable value
octant_analysis(mod)

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
